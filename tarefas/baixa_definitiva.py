import time

import pandas as pd
import pyperclip
from openpyxl import load_workbook, Workbook

from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import (
    NoSuchElementException,
    StaleElementReferenceException,
    TimeoutException,
)

from utilitarios import (
    caminho_arquivo,
    clicar_js,
    criar_wait,
    extrair_numero_processo,
    trocar_aba,
    trocar_para_iframe,
)

# ============================================================
# TAREFA: VERIFICAR BAIXA DEFINITIVA
# ============================================================

class ProcBaixa:
    """Verifica se os processos possuem movimento de baixa definitiva."""

    def executa(self, navegador):
        wait = criar_wait(navegador)

        processos = self.manipula_planilha()
        self.abre_pesquisa(wait, navegador)

        for numero in processos:
            print(f"Pesquisando processo {numero}")
            self.pesquisa_processo(
                wait,
                navegador,
                str(numero)
            )

        return "Tarefa 'Verificar baixa definitiva' concluída."

    @staticmethod
    def manipula_planilha():
        caminho = caminho_arquivo("analise.xlsx")

        if not caminho.exists():
            raise FileNotFoundError(
                f"Planilha não encontrada: {caminho}"
            )

        df = pd.read_excel(caminho)

        if "numero" not in df.columns:
            raise KeyError(
                "A coluna 'numero' não existe em analise.xlsx."
            )

        return df["numero"].dropna().astype(str)

    @staticmethod
    def abre_pesquisa(wait, navegador) -> None:
        xpaths = [
            "/html/body/nav/div/div[1]/ul/li/a/span",
            "/html/body/div[5]/div/nav/div[2]/ul/li[2]/a",
            (
                "/html/body/div[5]/div/nav/div[2]/ul/"
                "li[2]/div/ul/li[6]/a"
            ),
            (
                "/html/body/div[5]/div/nav/div[2]/ul/"
                "li[2]/div/ul/li[6]/div/ul/li[1]/a"
            ),
        ]

        for xpath in xpaths:
            wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, xpath)
                )
            ).click()

        time.sleep(0.5)

    def pesquisa_processo(
        self,
        wait,
        navegador,
        numero: str,
    ) -> None:
        abas_antes = navegador.window_handles

        campo = wait.until(
            EC.presence_of_element_located(
                (
                    By.ID,
                    "fPP:numeroProcesso:numeroSequencial",
                )
            )
        )

        campo.clear()
        pyperclip.copy(numero)
        campo.send_keys(Keys.CONTROL, "v")

        wait.until(
            EC.element_to_be_clickable(
                (By.ID, "fPP:searchProcessos")
            )
        ).click()

        time.sleep(3)

        try:
            resultado = navegador.find_element(
                By.XPATH,
                (
                    '//*[starts-with(@id, "fPP:processosTable:") '
                    'and contains(@id, ":j_id489")]'
                )
            )
            resultado.click()

            tem_baixa = self.verifica_baixa(
                abas_antes,
                navegador
            )

            categoria = (
                "com_baixa"
                if tem_baixa
                else "sem_baixa"
            )

        except NoSuchElementException:
            categoria = "nao_encontrado"
            print(f"Processo {numero} não encontrado.")

        self.salvar_processo(numero, categoria)

    @staticmethod
    def verifica_baixa(abas_antes, navegador) -> bool:
        wait = criar_wait(navegador)

        wait.until(
            lambda driver: len(driver.window_handles) > len(abas_antes)
        )

        aba_autos = navegador.window_handles[-1]
        navegador.switch_to.window(aba_autos)

        elementos = wait.until(
            EC.presence_of_all_elements_located(
                (By.CLASS_NAME, "texto-movimento")
            )
        )

        tem_baixa = any(
            "baixa definitiva" in elemento.text.lower()
            for elemento in elementos
        )

        print(
            "Encontrado: BAIXA DEFINITIVA"
            if tem_baixa
            else "Baixa definitiva não encontrada."
        )

        navegador.close()
        navegador.switch_to.window(navegador.window_handles[0])

        return tem_baixa

    @staticmethod
    def salvar_processo(
        numero_processo: str,
        categoria: str,
    ) -> None:
        caminho = caminho_arquivo("processos.xlsx")

        colunas = {
            "sem_baixa": ("A", "Processos Sem Baixa"),
            "com_baixa": ("B", "Processos Com Baixa"),
            "nao_encontrado": (
                "C",
                "Processos Não Encontrados",
            ),
        }

        if categoria not in colunas:
            raise ValueError(
                f"Categoria inválida: {categoria}"
            )

        if caminho.exists():
            workbook = load_workbook(caminho)
            planilha = workbook.active
        else:
            workbook = Workbook()
            planilha = workbook.active

            for letra, cabecalho in colunas.values():
                planilha[f"{letra}1"] = cabecalho

        letra_coluna, _ = colunas[categoria]

        linha = 2
        while planilha[f"{letra_coluna}{linha}"].value is not None:
            linha += 1

        planilha[f"{letra_coluna}{linha}"] = numero_processo
        workbook.save(caminho)

        print(
            f"Processo {numero_processo} salvo na categoria "
            f"'{categoria}'."
        )