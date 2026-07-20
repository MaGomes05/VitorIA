import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path
import unicodedata

import pandas as pd
import pyperclip
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.common.exceptions import (
    NoSuchElementException,
    StaleElementReferenceException,
    TimeoutException,
)
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


# ============================================================
# CONFIGURAÇÕES GERAIS
# ============================================================

URL_PJE = "https://pje.tre-pb.jus.br/pje/login.seam"
URL_DJE = "https://dje-consulta.tse.jus.br/#/dje/calendario?trib=TRE-PB"

TEMPO_ESPERA_PADRAO = 10
TEMPO_ESPERA_LOGIN = 20
TEMPO_CLICK = 0.7
TEMPO_TROCA_ABA = 0.7

DOCUMENTO_PROCURADO = "certidao de julgamento"

PADRAO_NUMERO_PROCESSO = (
    r"\d{7}-\d{2}\.\d{4}\.\d{1}\.\d{2}\.\d{4}"
)

PERFIS_GABINETES = {
    1: "GABJ01 - Gabinete Jurista 1 / Assessoria / Assessor Chefe",
    2: "GABJ02 - Gabinete Juiz de Direito 1 / Assessoria / Assessor Chefe",
    3: "GABJ03 - Gabinete Jurista 2 / Assessoria / Assessor Chefe",
    4: "GABJ04 - Gabinete Juiz de Direito 2 / Assessoria / Assessor Chefe",
    5: "GABJ05 - Gabinete Vice Presidência / Assessoria / Assessor Chefe",
    6: "GABJ06 - Gabinete Juiz Federal / Assessoria / Assessor Chefe",
}


# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================

def caminho_base() -> Path:
    """
    Retorna a pasta do programa.

    Funciona tanto durante a execução pelo Python quanto em um
    executável criado pelo PyInstaller.
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent

    return Path(__file__).resolve().parent


def caminho_arquivo(nome: str) -> Path:
    """Monta o caminho absoluto de um arquivo do projeto."""
    return caminho_base() / nome


def criar_wait(navegador, tempo=TEMPO_ESPERA_PADRAO) -> WebDriverWait:
    """Cria uma espera explícita para o navegador informado."""
    return WebDriverWait(navegador, tempo)


def clicar_js(
        navegador,
        elemento,
        pausa=TEMPO_CLICK,
        rolar=False
    ):
        """
        Clica em um elemento com uma pequena pausa.

        O scroll só acontece quando rolar=True.
        """
        if rolar:
            navegador.execute_script(
                """
                arguments[0].scrollIntoView({
                    block: 'center',
                    inline: 'center'
                });
                """,
                elemento
            )
            time.sleep(pausa)

        navegador.execute_script(
            "arguments[0].click();",
            elemento
        )

        time.sleep(pausa)

def normalizar_texto(texto: str) -> str:
    """
    Converte o texto para minúsculas e remove os acentos.

    Exemplo:
    'ACÓRDÃO' -> 'acordao'
    """
    texto = texto.lower().strip()

    return "".join(
        caractere
        for caractere in unicodedata.normalize("NFD", texto)
        if unicodedata.category(caractere) != "Mn"
    )

def trocar_aba(navegador, aba):
    navegador.switch_to.window(aba)
    time.sleep(TEMPO_TROCA_ABA)

def extrair_data(texto: str):
    """Extrai uma data no formato DD/MM/AAAA e retorna datetime."""
    correspondencia = re.search(r"\d{2}/\d{2}/\d{4}", texto)

    if not correspondencia:
        return None

    return datetime.strptime(correspondencia.group(), "%d/%m/%Y")


def extrair_numero_processo(texto: str):
    """Extrai o número CNJ de um processo."""
    correspondencia = re.search(PADRAO_NUMERO_PROCESSO, texto)
    return correspondencia.group() if correspondencia else None


def trocar_para_iframe(navegador, wait, iframe_id: str) -> None:
    """
    Retorna ao conteúdo principal e entra no iframe solicitado.
    """
    navegador.switch_to.default_content()
    wait.until(
        EC.frame_to_be_available_and_switch_to_it(
            (By.ID, iframe_id)
        )
    )


# ============================================================
# LOGIN
# ============================================================

class Inicial:
    """Responsável por iniciar o Firefox e efetuar o login no PJe."""

    def login(self, usuario: str, senha: str):
        servico = Service(
            executable_path=str(caminho_arquivo("geckodriver.exe"))
        )

        navegador = webdriver.Firefox(service=servico)
        wait = criar_wait(navegador, TEMPO_ESPERA_LOGIN)

        try:
            navegador.get(URL_PJE)

            try:
                wait.until(
                    EC.frame_to_be_available_and_switch_to_it(
                        (By.ID, "ssoFrame")
                    )
                )
                print("Entrou no iframe de login com sucesso.")
            except TimeoutException:
                print(
                    "Iframe 'ssoFrame' não encontrado. "
                    "Continuando no conteúdo principal."
                )

            campo_usuario = wait.until(
                EC.presence_of_element_located(
                    (By.ID, "username")
                )
            )
            campo_senha = wait.until(
                EC.presence_of_element_located(
                    (By.ID, "password")
                )
            )

            campo_usuario.clear()
            campo_usuario.send_keys(usuario)

            campo_senha.clear()
            campo_senha.send_keys(senha)

            navegador.find_element(
                By.XPATH,
                '//input[@value="Entrar"]'
            ).click()

            navegador.switch_to.default_content()

            # Aguarda o usuário preencher o 2FA manualmente.
            self.aguardar_conclusao_login(
                navegador,
                tempo_limite=300,
            )

            return navegador

        except Exception:
            # O navegador só é fechado quando o login não pôde ser concluído.
            navegador.quit()
            raise

    def aguardar_conclusao_login(
        self,
        navegador,
        tempo_limite: int = 300,
    ) -> bool:
        """
        Aguarda até o PJe carregar a tela principal.

        Durante esse período, o usuário pode preencher manualmente
        o código 2FA na janela do Firefox.
        """
        print(
            "Aguardando a conclusão do login e o preenchimento "
            "manual do código 2FA..."
        )

        def pagina_principal_carregada(driver):
            try:
                driver.switch_to.default_content()

                # Em alguns acessos o PJe exibe uma opção para pular
                # a verificação mobile. Se ela aparecer, tenta clicar.
                self._pular_verificacao_mobile(driver)

                # Elemento do menu de seleção de perfil.
                menu_perfil = driver.find_elements(
                    By.XPATH,
                    "/html/body/nav/div/div[2]/ul/li/a",
                )

                # Iframe principal do PJe.
                iframe_principal = driver.find_elements(
                    By.ID,
                    "ngFrame",
                )

                if menu_perfil or iframe_principal:
                    return True

                return False

            except Exception:
                return False

        try:
            WebDriverWait(
                navegador,
                tempo_limite,
                poll_frequency=1,
            ).until(pagina_principal_carregada)

            navegador.switch_to.default_content()

            print(
                "Login e autenticação 2FA concluídos. "
                "Tela principal do PJe identificada."
            )
            return True

        except TimeoutException as erro:
            raise TimeoutException(
                "O tempo para concluir o login e informar o "
                "código 2FA expirou."
            ) from erro

    @staticmethod
    def _pular_verificacao_mobile(navegador) -> None:
        xpath = (
            "/html/body/div[5]/div/div/div/div[2]/div/div/"
            "div/form/div/div[2]/div[4]/a"
        )

        try:
            elementos = navegador.find_elements(
                By.XPATH,
                xpath,
            )

            if elementos and elementos[0].is_displayed():
                elementos[0].click()
                print("Verificação mobile ignorada.")

        except Exception:
            # A ausência desse elemento é normal.
            pass


# ============================================================
# TAREFA: VISUALIZAR EXPEDIENTE DJE
# ============================================================

class Tarefa_visualizaDJE:
    """
    Automação da tarefa "Visualizar expediente DJE".

    Fluxo:
    1. Seleciona o perfil Coordenador de Processamento.
    2. Abre a tarefa Visualizar expediente DJE no PJe.
    3. Abre o DJE em uma segunda aba e mantém essa aba aberta.
    4. Para cada processo:
       - volta ao PJe;
       - abre o processo;
       - abre os autos em uma terceira aba;
       - consulta a data do expediente;
       - fecha somente a aba dos autos;
       - vai para a aba do DJE;
       - pesquisa o processo;
       - volta ao PJe;
       - finaliza ou mantém o processo na tarefa.
    """

    XPATH_MENU_PERFIL = "/html/body/nav/div/div[2]/ul/li/a"
    XPATH_CAMPO_PERFIL = (
        "/html/body/nav/div/div[2]/ul/li/div/form/div/"
        "div/table[1]/tbody/tr/td/input"
    )

    def __init__(self):
        self.aba_pje = None
        self.aba_dje = None

    def executa(self, navegador):
        wait = criar_wait(navegador)

        self.aba_pje = navegador.current_window_handle

        self.selecionar_perfil(wait, navegador)
        self.abrir_tarefa(wait, navegador)
        self.abrir_dje(wait, navegador)
        self.processar_processos(wait, navegador)

        return "Tarefa 'Visualizar expediente DJE' concluída."

    def selecionar_perfil(self, wait, navegador) -> None:
        trocar_aba(navegador, self.aba_pje)
        navegador.switch_to.default_content()

        wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, self.XPATH_MENU_PERFIL)
            )
        ).click()

        campo = wait.until(
            EC.visibility_of_element_located(
                (By.XPATH, self.XPATH_CAMPO_PERFIL)
            )
        )

        campo.clear()
        campo.send_keys("Coordenador de Processamento")

        xpath_link_perfil = (
            "/html/body/nav/div/div[2]/ul/li/div/form/div/"
            "div/table[2]/tbody/tr/td//a["
            "contains("
            "translate(normalize-space(.), "
            "'ABCDEFGHIJKLMNOPQRSTUVWXYZ', "
            "'abcdefghijklmnopqrstuvwxyz'), "
            "'coordenador de processamento'"
            ")"
            "]"
        )

        try:
            link_perfil = WebDriverWait(
                navegador,
                30,
                ignored_exceptions=(
                    StaleElementReferenceException,
                ),
            ).until(
                EC.element_to_be_clickable(
                    (By.XPATH, xpath_link_perfil)
                )
            )

            navegador.execute_script(
                "arguments[0].scrollIntoView({block: 'center'});",
                link_perfil,
            )

            link_perfil = navegador.find_element(
                By.XPATH,
                xpath_link_perfil,
            )

            navegador.execute_script(
                "arguments[0].click();",
                link_perfil,
            )

            print(
                "Perfil 'Coordenador de Processamento' selecionado."
            )

            navegador.switch_to.default_content()
            WebDriverWait(navegador, 30).until(
                EC.presence_of_element_located(
                    (By.ID, "ngFrame")
                )
            )

        except TimeoutException as erro:
            raise RuntimeError(
                "O perfil 'Coordenador de Processamento' "
                "não apareceu na lista."
            ) from erro

    def abrir_tarefa(self, wait, navegador) -> None:
        trocar_aba(navegador, self.aba_pje)
        trocar_para_iframe(navegador, wait, "ngFrame")

        indice = 1

        while True:
            xpath_tarefa = (
                "/html/body/app-root/selector/div/div/div[2]/"
                "right-panel/div/div/div[3]/tarefas/div/div[3]/"
                f"div[{indice}]/div/a/div/span[1]"
            )

            try:
                elemento = WebDriverWait(
                    navegador,
                    10,
                ).until(
                    EC.presence_of_element_located(
                        (By.XPATH, xpath_tarefa)
                    )
                )
            except TimeoutException as erro:
                raise RuntimeError(
                    "A tarefa 'Visualizar expediente DJE' "
                    "não foi encontrada."
                ) from erro

            texto_completo = elemento.text.strip()
            texto_sem_quantidade = re.match(
                r"^[^\d]*",
                texto_completo,
            ).group().strip()

            if texto_sem_quantidade == "Visualizar expediente DJE":
                elemento = wait.until(
                    EC.element_to_be_clickable(
                        (By.XPATH, xpath_tarefa)
                    )
                )

                clicar_js(
                    navegador,
                    elemento
                )

                print(
                    "Tarefa 'Visualizar expediente DJE' aberta."
                )

                time.sleep(2)
                return

            indice += 1

    def abrir_dje(self, wait, navegador) -> None:
        trocar_aba(navegador, self.aba_pje)
        abas_antes = set(navegador.window_handles)

        navegador.execute_script("window.open('about:blank', '_blank');")

        WebDriverWait(navegador, 15).until(
            lambda driver: len(set(driver.window_handles) - abas_antes) == 1
        )

        self.aba_dje = (
            set(navegador.window_handles) - abas_antes
        ).pop()

        trocar_aba(navegador, self.aba_dje)
        navegador.get(URL_DJE)

        WebDriverWait(navegador, 30).until(
            EC.presence_of_element_located(
                (By.ID, "mat-input-0")
            )
        )

        print("DJE aberto e mantido em uma aba separada.")

        trocar_aba(navegador, self.aba_pje)
        trocar_para_iframe(navegador, wait, "ngFrame")

    def processar_processos(self, wait, navegador) -> None:
        quantidade = self.obter_quantidade_processos(wait)

        if quantidade is None:
            raise RuntimeError(
                "A quantidade de processos não foi encontrada."
            )

        print(f"Quantidade de processos: {quantidade}")

        indice_lista = 1

        for numero_ordem in range(1, quantidade + 1):
            print(
                f"Processando {numero_ordem} de {quantidade}."
            )

            indice_lista = self.processar_um_processo(
                wait=wait,
                navegador=navegador,
                indice_lista=indice_lista,
            )

            time.sleep(2)

        print("Todos os processos foram percorridos.")

    @staticmethod
    def obter_quantidade_processos(wait):
        xpath_quantidade = (
            "/html/body/app-root/selector/div/div/div[2]/"
            "right-panel/div/processos-tarefa/div[1]/div[1]/"
            "filtro-tarefas/div/div[1]/div[2]/span"
        )

        elemento = wait.until(
            EC.presence_of_element_located(
                (By.XPATH, xpath_quantidade)
            )
        )

        correspondencia = re.search(r"\d+", elemento.text)
        return int(correspondencia.group()) if correspondencia else None

    @staticmethod
    def xpath_processo(indice: int) -> str:
        return (
            "/html/body/app-root/selector/div/div/div[2]/"
            "right-panel/div/processos-tarefa/div[1]/div[2]/"
            "div/div[1]/p-datalist/div/div/ul/"
            f"li[{indice}]/processo-datalist-card/div/div[3]/"
            "a/div/span[2]"
        )

    def processar_um_processo(
        self,
        wait,
        navegador,
        indice_lista: int,
    ) -> int:
        trocar_aba(navegador, self.aba_pje)
        trocar_para_iframe(navegador, wait, "ngFrame")

        xpath_processo = self.xpath_processo(indice_lista)

        elemento_processo = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, xpath_processo)
            )
        )

        numero_processo = extrair_numero_processo(
            elemento_processo.text
        )

        if not numero_processo:
            raise RuntimeError(
                "Número do processo não encontrado no cartão."
            )

        clicar_js(navegador, elemento_processo)
        print(f"Número do processo: {numero_processo}")

        data_expediente = self.obter_data_expediente(
            wait,
            navegador,
        )

        data_publicacao = self.pesquisar_no_dje(
            wait,
            navegador,
            numero_processo,
        )

        trocar_aba(navegador, self.aba_pje)

        deve_finalizar = self.compara_data(
            data_expediente,
            data_publicacao,
        )

        if deve_finalizar:
            self.finalizar_processo(wait, navegador)

            data_atual = datetime.now().strftime("%d-%m-%Y")
            self.log_processos(
                numero_processo,
                data_atual,
            )

            time.sleep(10)
            return indice_lista

        print(
            f"Processo {numero_processo} ainda não foi publicado."
        )
        return indice_lista + 1

    def obter_data_expediente(self, wait, navegador):
        trocar_aba(navegador, self.aba_pje)
        trocar_para_iframe(navegador, wait, "ngFrame")

        xpath_abrir_autos = (
            "/html/body/app-root/selector/div/div/div[2]/"
            "right-panel/div/processos-tarefa/div[2]/"
            "conteudo-tarefa/div[1]/div/div/div[2]/button[3]/i"
        )

        abas_antes = set(navegador.window_handles)

        botao_autos = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, xpath_abrir_autos)
            )
        )
        clicar_js(navegador, botao_autos)

        WebDriverWait(navegador, 20).until(
            lambda driver: len(
                set(driver.window_handles) - abas_antes
            ) == 1
        )

        aba_autos = (
            set(navegador.window_handles) - abas_antes
        ).pop()

        try:
            navegador.switch_to.window(aba_autos)

            botao_expedientes = WebDriverWait(
                navegador,
                30,
            ).until(
                EC.element_to_be_clickable(
                    (By.ID, "navbar:linkAbaExpedientes1")
                )
            )
            clicar_js(navegador, botao_expedientes)

            data_expediente = (
                self.procurar_data_diario_eletronico(navegador)
            )

            if not data_expediente:
                raise RuntimeError(
                    "Data do expediente no Diário Eletrônico "
                    "não encontrada."
                )

            print(
                "Data do expediente: "
                f"{data_expediente.strftime('%d/%m/%Y')}"
            )

            return data_expediente

        finally:
            if aba_autos in navegador.window_handles:
                navegador.switch_to.window(aba_autos)
                navegador.close()

            trocar_aba(navegador, self.aba_pje)

    @staticmethod
    def procurar_data_diario_eletronico(navegador):
        indice = 1

        while True:
            xpath_data = (
                "/html/body/div[1]/div[2]/div[2]/table/tbody/"
                "tr[2]/td/table/tbody/tr/td/div/div/div/div/"
                "div/div[2]/span/div/table/tbody/"
                f"tr[{indice}]/td[1]/span/div/span/div[2]"
            )

            try:
                elemento = WebDriverWait(
                    navegador,
                    5,
                ).until(
                    EC.presence_of_element_located(
                        (By.XPATH, xpath_data)
                    )
                )
            except TimeoutException:
                return None

            if "Diário Eletrônico" in elemento.text:
                return extrair_data(elemento.text)

            indice += 1

    def pesquisar_no_dje(
        self,
        wait,
        navegador,
        numero_processo: str,
    ):
        navegador.switch_to.window(self.aba_dje)

        data_default = datetime(2001, 5, 17)

        xpath_select_tribunal = (
            "/html/body/app-root/div/app-calendario/div/div[2]/"
            "app-pesquisa/app-pesquisa-form/div/div[1]/div[1]/"
            "mat-form-field/div/div[1]/div[3]"
        )
        xpath_tre_pb = (
            "/html/body/div[2]/div[2]/div/div/div/"
            "mat-option[17]/span"
        )
        xpath_pesquisar = (
            "/html/body/app-root/div/app-calendario/div/div[2]/"
            "app-pesquisa/app-pesquisa-form/div/div[5]/button[2]"
        )
        xpath_resultado_data = (
            "/html/body/app-root/div/app-calendario/div/div[8]/"
            "button/span"
        )

        try:
            WebDriverWait(navegador, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, xpath_select_tribunal)
                )
            ).click()

            WebDriverWait(navegador, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, xpath_tre_pb)
                )
            ).click()
        except TimeoutException:
            pass

        campo = WebDriverWait(
            navegador,
            20,
        ).until(
            EC.presence_of_element_located(
                (By.ID, "mat-input-0")
            )
        )

        campo.clear()
        pyperclip.copy(numero_processo)
        campo.send_keys(Keys.CONTROL, "v")

        WebDriverWait(navegador, 20).until(
            EC.element_to_be_clickable(
                (By.XPATH, xpath_pesquisar)
            )
        ).click()

        time.sleep(8)

        try:
            elemento_data = WebDriverWait(
                navegador,
                10,
            ).until(
                EC.presence_of_element_located(
                    (By.XPATH, xpath_resultado_data)
                )
            )
            data_publicacao = extrair_data(
                elemento_data.text
            )
        except TimeoutException:
            data_publicacao = None

        if data_publicacao:
            print(
                "Data da publicação no DJE: "
                f"{data_publicacao.strftime('%d/%m/%Y')}"
            )
            return data_publicacao

        print("Data da publicação não encontrada no DJE.")
        return data_default

    def finalizar_processo(self, wait, navegador) -> None:
        trocar_aba(navegador, self.aba_pje)
        trocar_para_iframe(navegador, wait, "ngFrame")

        botao_transicao = wait.until(
            EC.element_to_be_clickable(
                (By.ID, "btnTransicoesTarefa")
            )
        )
        clicar_js(navegador, botao_transicao)

        xpath_opcao = (
            "/html/body/app-root/selector/div/div/div[2]/"
            "right-panel/div/processos-tarefa/div[2]/"
            "conteudo-tarefa/div[1]/div/div/div[2]/"
            "div[2]/ul/li/a"
        )

        opcao = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, xpath_opcao)
            )
        )
        clicar_js(navegador, opcao)

    @staticmethod
    def compara_data(data_processo, data_publicacao) -> bool:
        return data_processo <= data_publicacao

    @staticmethod
    def log_processos(numero_processo: str, data_atual: str) -> None:
        nome_arquivo = f"processos_manipulados_{data_atual}.txt"
        caminho = caminho_arquivo(nome_arquivo)

        with caminho.open("a", encoding="utf-8") as arquivo:
            arquivo.write(f"{numero_processo}\n")

        print(
            f"Processo {numero_processo} salvo em {nome_arquivo}."
        )


# ============================================================
# TAREFA: ETIQUETA META 02 - 70%
# ============================================================

class Etiqueta_meta2_70:
    """Aplica a etiqueta 'META 02 - 70%' nos processos informados."""

    XPATH_MENU_PERFIL = "/html/body/nav/div/div[2]/ul/li/a"
    XPATH_CAMPO_PERFIL = (
        "/html/body/nav/div/div[2]/ul/li/div/form/div/"
        "div/table[1]/tbody/tr/td/input"
    )
    XPATH_RESULTADO_PERFIL = (
        "/html/body/nav/div/div[2]/ul/li/div/form/div/"
        "div/table[2]/tbody/tr/td[2]/a"
    )

    def executa(self, navegador):
        wait = criar_wait(navegador)

        for gabinete in range(1, 7):
            processos = self.manipula_planilha(gabinete)
            self.seleciona_gabinete(
                wait,
                navegador,
                gabinete
            )

            primeira_pesquisa = True

            for numero in processos:
                print(f"Pesquisando processo {numero}")

                primeira_pesquisa = self.pesquisa_processo(
                    wait,
                    navegador,
                    str(numero),
                    primeira_pesquisa,
                )

        return "Tarefa 'Etiqueta META 02 - 70%' concluída."

    def seleciona_gabinete(
        self,
        wait,
        navegador,
        gabinete: int,
    ) -> None:
        nome_perfil = PERFIS_GABINETES.get(gabinete)

        if not nome_perfil:
            raise ValueError(f"Gabinete inválido: {gabinete}")

        navegador.switch_to.default_content()

        wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, self.XPATH_MENU_PERFIL)
            )
        ).click()

        campo = wait.until(
            EC.presence_of_element_located(
                (By.XPATH, self.XPATH_CAMPO_PERFIL)
            )
        )

        campo.clear()
        campo.send_keys(nome_perfil)

        perfil = wait.until(
            EC.presence_of_element_located(
                (By.XPATH, self.XPATH_RESULTADO_PERFIL)
            )
        )

        if "Assessor Chefe" not in perfil.text:
            raise RuntimeError(
                f"Perfil do gabinete {gabinete} não encontrado."
            )

        perfil.click()
        print(f"Perfil do gabinete {gabinete} selecionado.")
        time.sleep(5)

    def pesquisa_processo(
        self,
        wait,
        navegador,
        numero: str,
        primeira_pesquisa: bool,
    ) -> bool:
        if primeira_pesquisa:
            trocar_para_iframe(
                navegador,
                wait,
                "ngFrame"
            )

            xpath_filtro = (
                "/html/body/app-root/selector/div/div/div[2]/"
                "right-panel/div/div/div[3]/tarefas/div/div[1]/div"
            )

            wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, xpath_filtro)
                )
            ).click()

        xpath_campo = (
            "/html/body/app-root/selector/div/div/div[2]/"
            "right-panel/div/div/div[3]/tarefas/div/div[2]/"
            "filtro-tarefas-pendentes/div/form/fieldset/"
            "div[1]/input"
        )
        xpath_botao_pesquisar = (
            "/html/body/app-root/selector/div/div/div[2]/"
            "right-panel/div/div/div[3]/tarefas/div/div[2]/"
            "filtro-tarefas-pendentes/div/form/fieldset/"
            "div[4]/button[1]"
        )

        campo = wait.until(
            EC.presence_of_element_located(
                (By.XPATH, xpath_campo)
            )
        )

        campo.clear()
        campo.send_keys(numero)

        wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, xpath_botao_pesquisar)
            )
        ).click()

        time.sleep(2)

        if self._abrir_resultado_pesquisa(navegador):
            self.etiqueta(wait, navegador)
            return True

        print(f"Processo {numero} não encontrado no gabinete.")
        return False

    @staticmethod
    def _abrir_resultado_pesquisa(navegador) -> bool:
        xpaths = [
            (
                "/html/body/app-root/selector/div/div/div[2]/"
                "right-panel/div/div/div[3]/tarefas/div/div[3]/"
                "div[1]/div/a/div/span[1]"
            ),
            (
                "/html/body/app-root/selector/div/div/div[2]/"
                "right-panel/div/div/div[3]/tarefas/div/div[3]/"
                "div/div/a/div/span[1]"
            ),
        ]

        for xpath in xpaths:
            try:
                navegador.find_element(By.XPATH, xpath).click()
                return True
            except NoSuchElementException:
                continue

        return False

    @staticmethod
    def etiqueta(wait, navegador) -> None:
        xpaths = {
            "selecionar_processo": (
                "/html/body/app-root/selector/div/div/div[2]/"
                "right-panel/div/processos-tarefa/div[1]/div[2]/"
                "div/div[1]/p-datalist/div/div/ul/li[1]/"
                "processo-datalist-card/div/div[2]/button/i"
            ),
            "abrir_etiquetas": (
                "/html/body/app-root/selector/div/div/div[2]/"
                "right-panel/div/processos-tarefa/div[1]/div[2]/"
                "div/div[1]/div[1]/acoes-processos-tarefa/"
                "div/div/div/button[2]/i"
            ),
            "adicionar_etiqueta": (
                "/html/body/app-root/selector/div/div/div[2]/"
                "right-panel/div/processos-tarefa/div[3]/"
                "etiquetar-lote/div/div/div/div[2]/div/"
                "pje-selecionar-etiquetas/div/div/table/"
                "tbody/tr/td[1]/button/i"
            ),
            "confirmar": (
                "/html/body/app-root/selector/div/div/div[2]/"
                "right-panel/div/processos-tarefa/div[3]/"
                "etiquetar-lote/div/div/div/div[3]/div/"
                "button[1]/span"
            ),
            "fechar": (
                "/html/body/app-root/selector/div/div/div[2]/"
                "right-panel/div/processos-tarefa/div[3]/"
                "etiquetar-lote/div/div/div/div[1]/button/span"
            ),
            "voltar": (
                "/html/body/app-root/selector/div/div/div[1]/"
                "side-bar/nav/ul/li[1]/a/i"
            ),
        }

        for chave in ("selecionar_processo", "abrir_etiquetas"):
            wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, xpaths[chave])
                )
            ).click()

        campo_etiqueta = wait.until(
            EC.presence_of_element_located(
                (By.ID, "itPesquisarEtiquetas")
            )
        )
        campo_etiqueta.clear()
        campo_etiqueta.send_keys("META 02 - 70%")

        for chave in (
            "adicionar_etiqueta",
            "confirmar",
            "fechar",
            "voltar",
        ):
            wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, xpaths[chave])
                )
            ).click()

            if chave == "confirmar":
                time.sleep(5)
            else:
                time.sleep(1)

    @staticmethod
    def manipula_planilha(gabinete: int):
        nome_planilha = f"Metas 2024_GABJ_0{gabinete}_Out.xlsx"
        caminho = caminho_arquivo(nome_planilha)

        if not caminho.exists():
            raise FileNotFoundError(
                f"Planilha não encontrada: {caminho}"
            )

        df = pd.read_excel(
            caminho,
            sheet_name="Meta 02 - 70%"
        )

        if "numero" not in df.columns:
            raise KeyError(
                f"A coluna 'numero' não existe em {nome_planilha}."
            )

        return df["numero"].dropna().astype(str)


# ============================================================
# TAREFA: VERIFICAR BAIXA DEFINITIVA
# ============================================================

class ProcBaixa:
    """Verifica se os processos possuem movimento de baixa definitiva."""

    def executa(self, navegador):
        wait = criar_wait(navegador)

        processos = self.manipula_planilha()
        self.abre_pesquisa(wait, navegador)

        for numero in processos:
            print(f"Pesquisando processo {numero}")
            self.pesquisa_processo(
                wait,
                navegador,
                str(numero)
            )

        return "Tarefa 'Verificar baixa definitiva' concluída."

    @staticmethod
    def manipula_planilha():
        caminho = caminho_arquivo("analise.xlsx")

        if not caminho.exists():
            raise FileNotFoundError(
                f"Planilha não encontrada: {caminho}"
            )

        df = pd.read_excel(caminho)

        if "numero" not in df.columns:
            raise KeyError(
                "A coluna 'numero' não existe em analise.xlsx."
            )

        return df["numero"].dropna().astype(str)

    @staticmethod
    def abre_pesquisa(wait, navegador) -> None:
        xpaths = [
            "/html/body/nav/div/div[1]/ul/li/a/span",
            "/html/body/div[5]/div/nav/div[2]/ul/li[2]/a",
            (
                "/html/body/div[5]/div/nav/div[2]/ul/"
                "li[2]/div/ul/li[6]/a"
            ),
            (
                "/html/body/div[5]/div/nav/div[2]/ul/"
                "li[2]/div/ul/li[6]/div/ul/li[1]/a"
            ),
        ]

        for xpath in xpaths:
            wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, xpath)
                )
            ).click()

        time.sleep(0.5)

    def pesquisa_processo(
        self,
        wait,
        navegador,
        numero: str,
    ) -> None:
        abas_antes = navegador.window_handles

        campo = wait.until(
            EC.presence_of_element_located(
                (
                    By.ID,
                    "fPP:numeroProcesso:numeroSequencial",
                )
            )
        )

        campo.clear()
        pyperclip.copy(numero)
        campo.send_keys(Keys.CONTROL, "v")

        wait.until(
            EC.element_to_be_clickable(
                (By.ID, "fPP:searchProcessos")
            )
        ).click()

        time.sleep(3)

        try:
            resultado = navegador.find_element(
                By.XPATH,
                (
                    '//*[starts-with(@id, "fPP:processosTable:") '
                    'and contains(@id, ":j_id489")]'
                )
            )
            resultado.click()

            tem_baixa = self.verifica_baixa(
                abas_antes,
                navegador
            )

            categoria = (
                "com_baixa"
                if tem_baixa
                else "sem_baixa"
            )

        except NoSuchElementException:
            categoria = "nao_encontrado"
            print(f"Processo {numero} não encontrado.")

        self.salvar_processo(numero, categoria)

    @staticmethod
    def verifica_baixa(abas_antes, navegador) -> bool:
        wait = criar_wait(navegador)

        wait.until(
            lambda driver: len(driver.window_handles) > len(abas_antes)
        )

        aba_autos = navegador.window_handles[-1]
        navegador.switch_to.window(aba_autos)

        elementos = wait.until(
            EC.presence_of_all_elements_located(
                (By.CLASS_NAME, "texto-movimento")
            )
        )

        tem_baixa = any(
            "baixa definitiva" in elemento.text.lower()
            for elemento in elementos
        )

        print(
            "Encontrado: BAIXA DEFINITIVA"
            if tem_baixa
            else "Baixa definitiva não encontrada."
        )

        navegador.close()
        navegador.switch_to.window(navegador.window_handles[0])

        return tem_baixa

    @staticmethod
    def salvar_processo(
        numero_processo: str,
        categoria: str,
    ) -> None:
        caminho = caminho_arquivo("processos.xlsx")

        colunas = {
            "sem_baixa": ("A", "Processos Sem Baixa"),
            "com_baixa": ("B", "Processos Com Baixa"),
            "nao_encontrado": (
                "C",
                "Processos Não Encontrados",
            ),
        }

        if categoria not in colunas:
            raise ValueError(
                f"Categoria inválida: {categoria}"
            )

        if caminho.exists():
            workbook = load_workbook(caminho)
            planilha = workbook.active
        else:
            workbook = Workbook()
            planilha = workbook.active

            for letra, cabecalho in colunas.values():
                planilha[f"{letra}1"] = cabecalho

        letra_coluna, _ = colunas[categoria]

        linha = 2
        while planilha[f"{letra_coluna}{linha}"].value is not None:
            linha += 1

        planilha[f"{letra_coluna}{linha}"] = numero_processo
        workbook.save(caminho)

        print(
            f"Processo {numero_processo} salvo na categoria "
            f"'{categoria}'."
        )


# ============================================================
# TAREFA: FINALIZAR PROCESSOS JULGADOS
# ============================================================

class ProcJulgado:
    """
    Automação da tarefa "Processo julgado".

    Fluxo:
    1. Seleciona o perfil Assessor-Chefe de Plenário.
    2. Abre a tarefa Processo julgado no PJe.
    3. Para cada processo:
       - abre o processo;
       - abre os autos em uma segunda aba;
       - consulta se tem acórdão;
       - fecha somente a aba dos autos;
       - volta para a tarfea no PJe;
       - finaliza ou mantém o processo na tarefa.
    """

    XPATH_MENU_PERFIL = "/html/body/nav/div/div[2]/ul/li/a"
    XPATH_CAMPO_PERFIL = (
        "/html/body/nav/div/div[2]/ul/li/div/form/div/"
        "div/table[1]/tbody/tr/td/input"
    )

    def __init__(self):
        self.aba_pje = None

    def executa(self, navegador):
        wait = criar_wait(navegador)

        self.aba_pje = navegador.current_window_handle

        self.selecionar_perfil(wait, navegador)
        self.abrir_tarefa(wait, navegador)
        self.processar_processos(wait, navegador)

        return "Tarefa 'Processo julgado' concluída."

    def selecionar_perfil(self, wait, navegador) -> None:
        trocar_aba(navegador, self.aba_pje)
        navegador.switch_to.default_content()

        # Abre o seletor de perfis
        botao_perfis = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, self.XPATH_MENU_PERFIL)
            )
        )
        clicar_js(navegador, botao_perfis)

        # Pesquisa pelo perfil
        campo = wait.until(
            EC.visibility_of_element_located(
                (By.XPATH, self.XPATH_CAMPO_PERFIL)
            )
        )

        campo.clear()
        campo.send_keys("Assessor-Chefe de Plenário")

        # Aguarda a tabela de resultados ser atualizada
        xpath_linhas = (
            "/html/body/nav/div/div[2]/ul/li/div/form/div/"
            "div/table[2]/tbody/tr"
        )

        try:
            WebDriverWait(
                navegador,
                30,
                ignored_exceptions=(
                    StaleElementReferenceException,
                ),
            ).until(
                EC.presence_of_all_elements_located(
                    (By.XPATH, xpath_linhas)
                )
            )

            time.sleep(1)

            linhas = navegador.find_elements(
                By.XPATH,
                xpath_linhas,
            )

            for linha in linhas:
                texto_linha = linha.text.strip().lower()

                print(f"Perfil encontrado: {linha.text.strip()}")

                if (
                    "assessor-chefe" in texto_linha
                    and "plenário" in texto_linha
                ):
                    # Procura o link na célula do nome do perfil,
                    # evitando clicar na estrela de favorito.
                    links_perfil = linha.find_elements(
                        By.XPATH,
                        "./td[2]//a"
                    )

                    if not links_perfil:
                        continue

                    # Localiza novamente a linha para evitar elemento obsoleto
                    linhas_atualizadas = navegador.find_elements(
                        By.XPATH,
                        xpath_linhas,
                    )

                    for linha_atualizada in linhas_atualizadas:
                        texto_atualizado = (
                            linha_atualizada.text.strip().lower()
                        )

                        if (
                            "assessor-chefe" in texto_atualizado
                            and "plenário" in texto_atualizado
                        ):
                            link_perfil = linha_atualizada.find_element(
                                By.XPATH,
                                "./td[2]//a"
                            )

                            clicar_js(
                                navegador,
                                link_perfil,
                            )

                            print(
                                "Perfil 'Assessor-Chefe de Plenário' "
                                "selecionado."
                            )

                            navegador.switch_to.default_content()

                            WebDriverWait(
                                navegador,
                                30,
                            ).until(
                                EC.presence_of_element_located(
                                    (By.ID, "ngFrame")
                                )
                            )

                            return

            raise RuntimeError(
                "O perfil 'Assessor-Chefe de Plenário' "
                "não foi encontrado entre os resultados."
            )

        except TimeoutException as erro:
            raise RuntimeError(
                "A lista de perfis não apareceu após a pesquisa."
            ) from erro

    def abrir_tarefa(self, wait, navegador) -> None:
        trocar_aba(navegador, self.aba_pje)
        trocar_para_iframe(navegador, wait, "ngFrame")

        indice = 1

        while True:
            xpath_tarefa = (
                "/html/body/app-root/selector/div/div/div[2]/"
                "right-panel/div/div/div[3]/tarefas/div/div[3]/"
                f"div[{indice}]/div/a/div/span[1]"
            )

            try:
                elemento = WebDriverWait(
                    navegador,
                    10,
                ).until(
                    EC.presence_of_element_located(
                        (By.XPATH, xpath_tarefa)
                    )
                )
            except TimeoutException as erro:
                raise RuntimeError(
                    "A tarefa 'Processo julgado' "
                    "não foi encontrada."
                ) from erro

            texto_completo = elemento.text.strip()
            texto_sem_quantidade = re.match(
                r"^[^\d]*",
                texto_completo,
            ).group().strip()

            if texto_sem_quantidade == "Processo julgado":
                elemento = wait.until(
                    EC.element_to_be_clickable(
                        (By.XPATH, xpath_tarefa)
                    )
                )

                clicar_js(
                    navegador,
                    elemento
                )

                print(
                    "Tarefa 'Processo julgado' aberta."
                )

                time.sleep(2)
                return

            indice += 1

    def processar_processos(self, wait, navegador) -> None:
        quantidade = self.obter_quantidade_processos(wait)

        if quantidade is None:
            raise RuntimeError(
                "A quantidade de processos não foi encontrada."
            )

        print(f"Quantidade de processos: {quantidade}")

        indice_lista = 1

        for numero_ordem in range(1, quantidade + 1):
            print(
                f"Processando {numero_ordem} de {quantidade}."
            )

            indice_lista = self.processar_um_processo(
                wait=wait,
                navegador=navegador,
                indice_lista=indice_lista,
            )

            time.sleep(2)

        print("Todos os processos foram percorridos.")

    @staticmethod
    def obter_quantidade_processos(wait):
        xpath_quantidade = (
            "/html/body/app-root/selector/div/div/div[2]/"
            "right-panel/div/processos-tarefa/div[1]/div[1]/"
            "filtro-tarefas/div/div[1]/div[2]/span"
        )

        elemento = wait.until(
            EC.presence_of_element_located(
                (By.XPATH, xpath_quantidade)
            )
        )

        correspondencia = re.search(r"\d+", elemento.text)
        return int(correspondencia.group()) if correspondencia else None

    @staticmethod
    def xpath_processo(indice: int) -> str:
        return (
            "/html/body/app-root/selector/div/div/div[2]/"
            "right-panel/div/processos-tarefa/div[1]/div[2]/"
            "div/div[1]/p-datalist/div/div/ul/"
            f"li[{indice}]/processo-datalist-card/div/div[3]/"
            "a/div/span[2]"
        )

    def processar_um_processo(
        self,
        wait,
        navegador,
        indice_lista: int,
    ) -> int:
        # Garante que estamos na aba principal do PJe
        trocar_aba(navegador, self.aba_pje)
        trocar_para_iframe(
            navegador,
            wait,
            "ngFrame"
        )

        xpath_processo = self.xpath_processo(
            indice_lista
        )

        elemento_processo = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, xpath_processo)
            )
        )

        numero_processo = extrair_numero_processo(
            elemento_processo.text
        )

        if not numero_processo:
            raise RuntimeError(
                "Número do processo não encontrado no cartão."
            )

        clicar_js(
            navegador,
            elemento_processo
        )

        print(
            f"Número do processo: {numero_processo}"
        )

        # Abre os autos e verifica se existe acórdão
        existe_certidao_julgamento = self.verifica_certidao_julgamento(
            wait,
            navegador,
        )

        # Depois da consulta, volta à aba principal do PJe
        trocar_aba(
            navegador,
            self.aba_pje
        )

        if existe_certidao_julgamento:
            self.finalizar_processo(
                wait,
                navegador
            )

            data_atual = datetime.now().strftime(
                "%d-%m-%Y"
            )

            self.log_processos(
                numero_processo,
                data_atual,
            )

            # Como o processo sai da lista, mantém o mesmo índice
            time.sleep(10)
            return indice_lista

        print(
            f"Processo {numero_processo} ainda não possui acórdão."
        )

        # Como o processo continua na lista, avança o índice
        return indice_lista + 1
    
    def verifica_certidao_julgamento(
        self,
        wait,
        navegador,
    ) -> bool:
        """
        Abre os autos, pesquisa por acórdão no texto visível da página,
        fecha somente a aba dos autos e volta à aba principal do PJe.
        """

        trocar_aba(
            navegador,
            self.aba_pje
        )

        trocar_para_iframe(
            navegador,
            wait,
            "ngFrame"
        )

        xpath_abrir_autos = (
            "/html/body/app-root/selector/div/div/div[2]/"
            "right-panel/div/processos-tarefa/div[2]/"
            "conteudo-tarefa/div[1]/div/div/div[2]/"
            "button[3]/i"
        )

        abas_antes = set(
            navegador.window_handles
        )

        botao_autos = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, xpath_abrir_autos)
            )
        )

        clicar_js(
            navegador,
            botao_autos
        )

        WebDriverWait(
            navegador,
            30,
        ).until(
            lambda driver: len(
                set(driver.window_handles) - abas_antes
            ) == 1
        )

        abas_novas = (
            set(navegador.window_handles) - abas_antes
        )

        aba_autos = abas_novas.pop()

        try:
            trocar_aba(
                navegador,
                aba_autos
            )

            navegador.switch_to.default_content()

            WebDriverWait(
                navegador,
                30,
            ).until(
                EC.presence_of_element_located(
                    (By.TAG_NAME, "body")
                )
            )

            # Dá tempo para a árvore/documentos serem carregados.
            time.sleep(3)

            existe_certidao_julgamento = self.procurar_certidao_julgamento_pagina(
                navegador
            )

            if existe_certidao_julgamento:
                print("Encontrado: Certidão de Julgamento")
            else:
                print("Certidão de Julgamento não encontrada.")

            return existe_certidao_julgamento

        finally:
            if aba_autos in navegador.window_handles:
                trocar_aba(
                    navegador,
                    aba_autos
                )
                navegador.close()

            trocar_aba(
                navegador,
                self.aba_pje
            )
    
    def procurar_certidao_julgamento_pagina(
        self,
        navegador,
    ) -> bool:
        """
        Procura o documento 'certidão de julgamento' no conteúdo principal e nos iframes.
        """

        navegador.switch_to.default_content()

        # Primeiro verifica o conteúdo principal da página.
        texto_principal = navegador.execute_script(
            "return document.body ? document.body.innerText : '';"
        )

        texto_normalizado = normalizar_texto(
            texto_principal or ""
        )

        print(
            "Texto principal contém 'certidao de julgamento':",
            DOCUMENTO_PROCURADO in texto_normalizado
        )

        if DOCUMENTO_PROCURADO in texto_normalizado:
            self.imprimir_trechos_certidao_julgamento(
                texto_principal
            )
            return True

        # Depois verifica possíveis iframes.
        iframes = navegador.find_elements(
            By.TAG_NAME,
            "iframe"
        )

        print(
            f"Quantidade de iframes nos autos: {len(iframes)}"
        )

        for indice in range(len(iframes)):
            try:
                navegador.switch_to.default_content()

                # Localiza novamente para evitar elemento obsoleto.
                iframes_atualizados = navegador.find_elements(
                    By.TAG_NAME,
                    "iframe"
                )

                navegador.switch_to.frame(
                    iframes_atualizados[indice]
                )

                texto_iframe = navegador.execute_script(
                    "return document.body ? document.body.innerText : '';"
                )

                texto_iframe_normalizado = normalizar_texto(
                    texto_iframe or ""
                )

                print(
                    f"Iframe {indice} contém 'certidao de julgamento':",
                    DOCUMENTO_PROCURADO in texto_iframe_normalizado
                )

                if DOCUMENTO_PROCURADO in texto_iframe_normalizado:
                    self.imprimir_trechos_certidao_julgamento(
                        texto_iframe
                    )

                    navegador.switch_to.default_content()
                    return True

            except (
                StaleElementReferenceException,
                NoSuchElementException,
            ):
                continue

        navegador.switch_to.default_content()
        return False
    
    @staticmethod
    def imprimir_trechos_certidao_julgamento(texto: str) -> None:
        """
        Mostra no terminal as linhas em que aparece a palavra acórdão.
        """

        for linha in texto.splitlines():
            if DOCUMENTO_PROCURADO in normalizar_texto(linha):
                print(
                    f"Trecho encontrado: {linha.strip()}"
                )
    
    def finalizar_processo(self, wait, navegador) -> None:
        trocar_aba(navegador, self.aba_pje)
        trocar_para_iframe(navegador, wait, "ngFrame")

        botao_transicao = wait.until(
            EC.element_to_be_clickable(
                (By.ID, "btnTransicoesTarefa")
            )
        )
        clicar_js(navegador, botao_transicao)

        xpath_opcao = (
            "/html/body/app-root/selector/div/div/div[2]/"
            "right-panel/div/processos-tarefa/div[2]/"
            "conteudo-tarefa/div[1]/div/div/div[2]/"
            "div[2]/ul/li/a"
        )

        opcao = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, xpath_opcao)
            )
        )
        clicar_js(navegador, opcao)

    @staticmethod
    def log_processos(numero_processo: str, data_atual: str) -> None:
        nome_arquivo = f"processos_julgados_finalizados_{data_atual}.txt"
        caminho = caminho_arquivo(nome_arquivo)

        with caminho.open("a", encoding="utf-8") as arquivo:
            arquivo.write(f"{numero_processo}\n")

        print(
            f"Processo {numero_processo} salvo em {nome_arquivo}."
        )